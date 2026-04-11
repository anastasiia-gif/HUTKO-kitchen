"""
HUTKO — shop_data.py
Serves product/bundle data, images, and site settings from hutko_shop.xlsx.

Endpoints:
  GET  /api/shop/all       — products + bundles
  GET  /api/shop/products  — products only
  GET  /api/shop/bundles   — bundles only
  GET  /api/shop/settings  — contact info & business settings
  GET  /api/shop/debug     — admin debug (requires X-Admin-Password header)
  POST /api/admin/upload-excel — replace the Excel file (requires X-Admin-Password)
"""

import os
import traceback
from flask import Blueprint, jsonify, request
from openpyxl import load_workbook

shop_bp    = Blueprint('shop', __name__)
EXCEL_PATH = os.environ.get('SHOP_EXCEL_PATH', 'hutko_shop.xlsx')
_cache     = {}


# ── PARSER ───────────────────────────────────────────────────────────

def _load_excel():
    if not os.path.exists(EXCEL_PATH):
        print(f"[SHOP] Excel not found at: {os.path.abspath(EXCEL_PATH)}")
        return {'products': [], 'bundles': [], 'images': {}, 'settings': {}}

    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    except Exception as e:
        print(f"[SHOP] Failed to open Excel: {e}")
        return {'products': [], 'bundles': [], 'images': {}, 'settings': {}}

    print(f"[SHOP] Sheets: {wb.sheetnames}")

    # ── 1. IMAGES sheet — keyed by item_id ───────────────
    images = {}   # { 'syrnyky': 'assets/products/syrnyky.png', ... }
    if 'Images' in wb.sheetnames:
        ws   = wb['Images']
        rows = list(ws.iter_rows(values_only=True))
        hdr  = None
        for i, row in enumerate(rows):
            if row and row[0] == 'type':
                hdr = i; break
        if hdr is not None:
            for row in rows[hdr + 1:]:
                if not row or not any(row): continue
                item_id   = str(row[1] or '').strip()
                photo_file = str(row[3] or '').strip()
                if item_id and photo_file:
                    images[item_id] = photo_file
        print(f"[SHOP] Images loaded: {len(images)} entries")

    # ── 2. SETTINGS sheet ────────────────────────────────
    settings = {}
    if 'Settings' in wb.sheetnames:
        ws   = wb['Settings']
        rows = list(ws.iter_rows(values_only=True))
        hdr  = None
        for i, row in enumerate(rows):
            if row and row[0] == 'setting':
                hdr = i; break
        if hdr is not None:
            for row in rows[hdr + 1:]:
                if not row or not any(row): continue
                key = str(row[0] or '').strip()
                val = str(row[1] or '').strip() if row[1] is not None else ''
                if key:
                    settings[key] = val
        print(f"[SHOP] Settings loaded: {len(settings)} keys")

    # ── 3. PRODUCTS sheet ────────────────────────────────
    products = []
    if 'Products' in wb.sheetnames:
        ws   = wb['Products']
        rows = list(ws.iter_rows(values_only=True))

        hdr = None
        for i, row in enumerate(rows):
            if row and row[1] == 'id':
                hdr = i; break

        if hdr is not None:
            headers = [str(c).strip() if c else '' for c in rows[hdr][1:]]
            for row in rows[hdr + 1:]:
                if row and row[1] == 'product_id': break  # hit variants table
                vals = row[1:] if row else []
                if not any(vals): continue
                d = dict(zip(headers, vals))
                if not d.get('id'): continue
                if str(d.get('active', 'yes')).lower() != 'yes': continue

                pid   = str(d['id'])
                photo = images.get(pid) or str(d.get('photo_file', '') or '')

                products.append({
                    'id':        pid,
                    'name_en':   str(d.get('name_en', '') or ''),
                    'name_ua':   str(d.get('name_ua', '') or ''),
                    'name_nl':   str(d.get('name_nl', '') or ''),
                    'category':  str(d.get('category', '') or ''),
                    'desc_en':   str(d.get('description_en', '') or ''),
                    'desc_ua':   str(d.get('description_ua', '') or ''),
                    'desc_nl':   str(d.get('description_nl', '') or ''),
                    'base_price': float(d['base_price']) if d.get('base_price') else 0,
                    'unit':      str(d.get('unit', '') or ''),
                    'badge':     str(d.get('badge', '') or ''),
                    'photo':     photo,
                    'dietary':   [t.strip() for t in str(d.get('dietary','') or '').split(',') if t.strip()],
                })

        print(f"[SHOP] Products loaded: {len(products)}")

        # Variants
        variants_by_id = {}
        var_hdr = None
        for i, row in enumerate(rows):
            if row and row[1] == 'product_id':
                var_hdr = i; break

        if var_hdr is not None:
            var_headers = [str(c).strip() if c else '' for c in rows[var_hdr][1:]]
            for row in rows[var_hdr + 1:]:
                vals = row[1:] if row else []
                if not any(vals): continue
                d   = dict(zip(var_headers, vals))
                pid = str(d.get('product_id', ''))
                if pid and str(d.get('active', 'yes')).lower() == 'yes':
                    variants_by_id.setdefault(pid, []).append({
                        'label': str(d.get('label', '')),
                        'price': float(d['price']) if d.get('price') else 0,
                    })

        for p in products:
            p['variants'] = variants_by_id.get(p['id'], [
                {'label': p['unit'], 'price': p['base_price']}
            ])

    # ── 4. BUNDLES sheet ─────────────────────────────────
    bundles = []
    if 'Bundles' in wb.sheetnames:
        ws   = wb['Bundles']
        rows = list(ws.iter_rows(values_only=True))

        hdr = None
        for i, row in enumerate(rows):
            if row and row[1] == 'id':
                hdr = i; break

        if hdr is not None:
            headers = [str(c).strip() if c else '' for c in rows[hdr][1:]]
            for row in rows[hdr + 1:]:
                vals = row[1:] if row else []
                if not any(vals): continue
                d = dict(zip(headers, vals))
                if not d.get('id'): continue
                if str(d.get('active', 'yes')).lower() != 'yes': continue

                bid   = str(d['id'])
                photo = images.get(bid) or str(d.get('photo_file', '') or '')

                items = []
                for part in str(d.get('items', '') or '').split(','):
                    part = part.strip()
                    if ':' in part:
                        p_id, qty = part.split(':', 1)
                        try:
                            items.append({'product_id': p_id.strip(), 'qty': int(qty.strip())})
                        except ValueError:
                            pass

                bundles.append({
                    'id':             bid,
                    'name_en':        str(d.get('name_en', '') or ''),
                    'name_ua':        str(d.get('name_ua', '') or ''),
                    'name_nl':        str(d.get('name_nl', '') or ''),
                    'size_label':     str(d.get('size_label', '') or ''),
                    'items':          items,
                    'original_price': float(d['original_price']) if d.get('original_price') else 0,
                    'discount_price': float(d['discount_price']) if d.get('discount_price') else 0,
                    'photo':          photo,
                    'badge':          str(d.get('badge', '') or ''),
                })

        print(f"[SHOP] Bundles loaded: {len(bundles)}")

    wb.close()
    return {'products': products, 'bundles': bundles, 'images': images, 'settings': settings}


# ── CACHE ─────────────────────────────────────────────────────────────

def get_shop_data():
    global _cache
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
        if _cache.get('mtime') != mtime:
            print(f"[SHOP] Reloading Excel")
            data        = _load_excel()
            _cache      = data
            _cache['mtime'] = mtime
    except FileNotFoundError:
        print(f"[SHOP] Excel missing: {EXCEL_PATH}")
        if not _cache.get('products'):
            _cache = {'products': [], 'bundles': [], 'images': {}, 'settings': {}}
    except Exception as e:
        print(f"[SHOP] Error: {e}\n{traceback.format_exc()}")
        if not _cache.get('products'):
            _cache = {'products': [], 'bundles': [], 'images': {}, 'settings': {}}
    return _cache


# ── ENDPOINTS ─────────────────────────────────────────────────────────

@shop_bp.route('/api/shop/products', methods=['GET'])
def get_products():
    return jsonify({'products': get_shop_data().get('products', [])}), 200


@shop_bp.route('/api/shop/bundles', methods=['GET'])
def get_bundles():
    return jsonify({'bundles': get_shop_data().get('bundles', [])}), 200


@shop_bp.route('/api/shop/all', methods=['GET'])
def get_all():
    d = get_shop_data()
    return jsonify({'products': d.get('products', []), 'bundles': d.get('bundles', [])}), 200


@shop_bp.route('/api/shop/settings', methods=['GET'])
def get_settings():
    return jsonify({'settings': get_shop_data().get('settings', {})}), 200


@shop_bp.route('/api/shop/debug', methods=['GET'])
def debug_excel():
    if request.headers.get('X-Admin-Password','') != os.environ.get('ADMIN_PASSWORD',''):
        return jsonify({'error': 'Unauthorized'}), 401
    d = get_shop_data()
    return jsonify({
        'excel_path':     os.path.abspath(EXCEL_PATH),
        'excel_exists':   os.path.exists(EXCEL_PATH),
        'products_count': len(d.get('products', [])),
        'bundles_count':  len(d.get('bundles', [])),
        'images_count':   len(d.get('images', {})),
        'settings_keys':  list(d.get('settings', {}).keys()),
        'cached_mtime':   _cache.get('mtime'),
    }), 200


@shop_bp.route('/api/admin/upload-excel', methods=['POST'])
def upload_excel():
    if request.headers.get('X-Admin-Password','') != os.environ.get('ADMIN_PASSWORD',''):
        return jsonify({'error': 'Unauthorized'}), 401
    if 'file' not in request.files:
        return jsonify({'error': 'No file. Use field name "file".'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx files accepted.'}), 400
    try:
        f.save(EXCEL_PATH)
        global _cache
        _cache = {}
        d = get_shop_data()
        return jsonify({
            'message':        'Uploaded and parsed successfully.',
            'products_count': len(d.get('products', [])),
            'bundles_count':  len(d.get('bundles', [])),
            'settings_keys':  list(d.get('settings', {}).keys()),
        }), 200
    except Exception as e:
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500