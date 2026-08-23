"""
HUTKO — migrate_excel_to_db.py
One-off (re-runnable) importer: hutko_shop.xlsx  ->  database tables.

Run ONCE at cutover. After that, manage products in the admin panel — do NOT
re-run this against the old Excel or it will overwrite your admin edits with the
spreadsheet's contents (products/bundles are replaced by id). Settings from the
sheet are updated by key; business-rule defaults are only inserted if missing.

Usage:
    python migrate_excel_to_db.py [path/to/hutko_shop.xlsx]
"""

import os
import sys
import json
from openpyxl import load_workbook

from database import init_db, get_db, _placeholder
from settings_store import DEFAULTS

_p = _placeholder()


def _normalize_photo(path):
    if not path:
        return ''
    path = str(path).strip()
    dot = path.rfind('.')
    return path if dot == -1 else path[:dot] + path[dot:].lower()


def _find_header(rows, col_index, marker):
    for i, row in enumerate(rows):
        if row and len(row) > col_index and row[col_index] == marker:
            return i
    return None


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def parse_workbook(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames

    images = {}
    if 'Images' in sheets:
        rows = _rows(wb['Images'])
        hdr = _find_header(rows, 0, 'type')
        if hdr is not None:
            for row in rows[hdr + 1:]:
                if not row or not any(row):
                    continue
                item_id = str(row[1] or '').strip() if len(row) > 1 else ''
                photo   = str(row[3] or '').strip() if len(row) > 3 else ''
                if item_id and photo:
                    images[item_id] = photo

    settings = {}
    if 'Settings' in sheets:
        rows = _rows(wb['Settings'])
        hdr = _find_header(rows, 0, 'setting')
        if hdr is not None:
            for row in rows[hdr + 1:]:
                if not row or not any(row):
                    continue
                key = str(row[0] or '').strip()
                val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                if key:
                    settings[key] = val

    products, variants_by_id = [], {}
    if 'Products' in sheets:
        rows = _rows(wb['Products'])
        hdr = _find_header(rows, 1, 'id')
        if hdr is not None:
            headers = [str(c).strip() if c else '' for c in rows[hdr][1:]]
            for row in rows[hdr + 1:]:
                if row and len(row) > 1 and row[1] == 'product_id':
                    break
                vals = row[1:] if row else []
                if not any(vals):
                    continue
                d = dict(zip(headers, vals))
                if not d.get('id'):
                    continue
                pid = str(d['id'])
                raw_photo = images.get(pid) or str(d.get('photo_file', '') or '')
                gallery = [_normalize_photo(g) for g in str(d.get('gallery', '') or '').split(',') if g.strip()]
                dietary = [t.strip() for t in str(d.get('dietary', '') or '').split(',') if t.strip()]
                products.append({
                    'id': pid, 'category': str(d.get('category', '') or ''),
                    'name_en': str(d.get('name_en', '') or ''), 'name_ua': str(d.get('name_ua', '') or ''),
                    'name_nl': str(d.get('name_nl', '') or ''),
                    'desc_en': str(d.get('description_en', '') or ''), 'desc_ua': str(d.get('description_ua', '') or ''),
                    'desc_nl': str(d.get('description_nl', '') or ''),
                    'about_en': str(d.get('about_en', '') or ''), 'about_ua': str(d.get('about_ua', '') or ''),
                    'about_nl': str(d.get('about_nl', '') or ''),
                    'prepare_en': str(d.get('prepare_en', '') or ''), 'prepare_ua': str(d.get('prepare_ua', '') or ''),
                    'prepare_nl': str(d.get('prepare_nl', '') or ''),
                    'ingredients_en': str(d.get('ingredients_en', '') or ''), 'ingredients_ua': str(d.get('ingredients_ua', '') or ''),
                    'ingredients_nl': str(d.get('ingredients_nl', '') or ''),
                    'hutko_tip_en': str(d.get('hutko_tip_en', '') or ''), 'hutko_tip_ua': str(d.get('hutko_tip_ua', '') or ''),
                    'hutko_tip_nl': str(d.get('hutko_tip_nl', '') or ''),
                    'storage_en': str(d.get('storage_en', '') or ''), 'storage_ua': str(d.get('storage_ua', '') or ''),
                    'storage_nl': str(d.get('storage_nl', '') or ''),
                    'base_price': float(d['base_price']) if d.get('base_price') else 0,
                    'unit': str(d.get('unit', '') or ''), 'badge': str(d.get('badge', '') or ''),
                    'photo': _normalize_photo(raw_photo), 'gallery': gallery, 'dietary': dietary,
                    'active': 1 if str(d.get('active', 'yes')).lower() == 'yes' else 0,
                })
            vhdr = _find_header(rows, 1, 'product_id')
            if vhdr is not None:
                vheaders = [str(c).strip() if c else '' for c in rows[vhdr][1:]]
                for row in rows[vhdr + 1:]:
                    vals = row[1:] if row else []
                    if not any(vals):
                        continue
                    d = dict(zip(vheaders, vals))
                    pid = str(d.get('product_id', '') or '')
                    if not pid:
                        continue
                    variants_by_id.setdefault(pid, []).append({
                        'label': str(d.get('label', '') or ''),
                        'price': float(d['price']) if d.get('price') else 0,
                        'active': 1 if str(d.get('active', 'yes')).lower() == 'yes' else 0,
                    })

    bundles = []
    if 'Bundles' in sheets:
        rows = _rows(wb['Bundles'])
        hdr = _find_header(rows, 1, 'id')
        if hdr is not None:
            headers = [str(c).strip() if c else '' for c in rows[hdr][1:]]
            for row in rows[hdr + 1:]:
                vals = row[1:] if row else []
                if not any(vals):
                    continue
                d = dict(zip(headers, vals))
                if not d.get('id'):
                    continue
                bid = str(d['id'])
                items = []
                for part in str(d.get('items', '') or '').split(','):
                    part = part.strip()
                    if ':' in part:
                        p_id, qty = part.split(':', 1)
                        try:
                            items.append({'product_id': p_id.strip(), 'qty': int(qty.strip())})
                        except ValueError:
                            pass
                raw_photo = images.get(bid) or str(d.get('photo_file', '') or '')
                bundles.append({
                    'id': bid, 'name_en': str(d.get('name_en', '') or ''), 'name_ua': str(d.get('name_ua', '') or ''),
                    'name_nl': str(d.get('name_nl', '') or ''), 'size_label': str(d.get('size_label', '') or ''),
                    'items': items,
                    'original_price': float(d['original_price']) if d.get('original_price') else 0,
                    'discount_price': float(d['discount_price']) if d.get('discount_price') else 0,
                    'photo': _normalize_photo(raw_photo), 'badge': str(d.get('badge', '') or ''),
                    'choice_en': str(d.get('choice_en', '') or ''), 'choice_ua': str(d.get('choice_ua', '') or ''),
                    'choice_nl': str(d.get('choice_nl', '') or ''),
                    'active': 1 if str(d.get('active', 'yes')).lower() == 'yes' else 0,
                })

    wb.close()
    return {'products': products, 'variants': variants_by_id,
            'bundles': bundles, 'settings': settings, 'images': images}


def import_into_db(data):
    conn = get_db()
    cur = conn.cursor()

    for i, p in enumerate(data['products']):
        cols = ['id', 'category', 'name_en', 'name_ua', 'name_nl',
                'desc_en', 'desc_ua', 'desc_nl', 'about_en', 'about_ua', 'about_nl',
                'prepare_en', 'prepare_ua', 'prepare_nl',
                'ingredients_en', 'ingredients_ua', 'ingredients_nl',
                'hutko_tip_en', 'hutko_tip_ua', 'hutko_tip_nl',
                'storage_en', 'storage_ua', 'storage_nl',
                'base_price', 'unit', 'badge', 'photo', 'gallery', 'dietary',
                'active', 'sort_order']
        vals = [p['id'], p['category'], p['name_en'], p['name_ua'], p['name_nl'],
                p['desc_en'], p['desc_ua'], p['desc_nl'], p['about_en'], p['about_ua'], p['about_nl'],
                p['prepare_en'], p['prepare_ua'], p['prepare_nl'],
                p['ingredients_en'], p['ingredients_ua'], p['ingredients_nl'],
                p['hutko_tip_en'], p['hutko_tip_ua'], p['hutko_tip_nl'],
                p['storage_en'], p['storage_ua'], p['storage_nl'],
                p['base_price'], p['unit'], p['badge'], p['photo'],
                json.dumps(p['gallery']), json.dumps(p['dietary']), p['active'], i]
        cur.execute(f"INSERT OR REPLACE INTO products ({','.join(cols)}) VALUES ({','.join([_p]*len(cols))})", vals)
        cur.execute(f"DELETE FROM product_variants WHERE product_id={_p}", (p['id'],))
        for j, v in enumerate(data['variants'].get(p['id'], [])):
            cur.execute(
                f"INSERT INTO product_variants (product_id, label, price, active, sort_order) "
                f"VALUES ({_p},{_p},{_p},{_p},{_p})",
                (p['id'], v['label'], v['price'], v['active'], j))

    for i, b in enumerate(data['bundles']):
        cols = ['id', 'name_en', 'name_ua', 'name_nl', 'size_label', 'items',
                'original_price', 'discount_price', 'photo', 'badge',
                'choice_en', 'choice_ua', 'choice_nl', 'active', 'sort_order']
        vals = [b['id'], b['name_en'], b['name_ua'], b['name_nl'], b['size_label'],
                json.dumps(b['items']), b['original_price'], b['discount_price'],
                b['photo'], b['badge'], b['choice_en'], b['choice_ua'], b['choice_nl'],
                b['active'], i]
        cur.execute(f"INSERT OR REPLACE INTO bundles ({','.join(cols)}) VALUES ({','.join([_p]*len(cols))})", vals)

    for key, val in data['settings'].items():
        cur.execute(
            f"INSERT INTO settings (key, value, updated_at) VALUES ({_p},{_p},datetime('now')) "
            f"ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
            (key, val))

    for key, val in DEFAULTS.items():
        cur.execute(f"INSERT OR IGNORE INTO settings (key, value) VALUES ({_p},{_p})", (key, val))

    conn.commit()
    variant_count = sum(len(v) for v in data['variants'].values())
    summary = {'products': len(data['products']), 'variants': variant_count,
               'bundles': len(data['bundles']), 'settings_from_sheet': len(data['settings'])}
    conn.close()
    return summary


def run(path):
    if not os.path.exists(path):
        print(f"[MIGRATE] Excel not found: {os.path.abspath(path)}")
        return None
    init_db()
    data = parse_workbook(path)
    summary = import_into_db(data)
    print("[MIGRATE] Done ✓")
    print(f"  Products : {summary['products']}")
    print(f"  Variants : {summary['variants']}")
    print(f"  Bundles  : {summary['bundles']}")
    print(f"  Settings : {summary['settings_from_sheet']} from sheet (+ defaults seeded if missing)")
    return summary


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else os.environ.get('SHOP_EXCEL_PATH', 'hutko_shop.xlsx')
    run(xlsx)
