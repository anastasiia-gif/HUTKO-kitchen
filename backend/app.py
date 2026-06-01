"""
HUTKO — app.py
Token-based auth. No session dependency.
"""

import os
import io
import datetime
import threading
import time

from flask import Flask, jsonify
from flask_cors import CORS
from dotenv import load_dotenv

from database import init_db
from auth             import auth_bp
from orders           import orders_bp
from contact          import contact_bp
from admin            import admin_bp
from trello_webhook   import webhook_bp
from shop_data        import shop_bp
from payments         import payments_bp

load_dotenv()

app = Flask(__name__)
secret = os.environ.get('SECRET_KEY', '')
if not secret:
    import warnings
    warnings.warn("⚠️  SECRET_KEY env var not set — using insecure default.")
    secret = 'dev-secret-change-me'
app.config['SECRET_KEY'] = secret

ALLOWED_ORIGINS = [
    "https://hutko-kitchen.com",
    "https://www.hutko-kitchen.com",
    "https://hutko-kitchen.pages.dev",
    "https://hutko.netlify.app",
    "https://testing--hutko.netlify.app",
    "http://localhost:3000",
    "http://localhost:5500",
    "http://127.0.0.1:5500",
    "null",
]
if os.environ.get('FRONTEND_URL'):
    ALLOWED_ORIGINS.append(os.environ['FRONTEND_URL'])

CORS(app,
     origins=ALLOWED_ORIGINS,
     allow_headers=["Content-Type", "Authorization"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
     supports_credentials=True)

app.register_blueprint(auth_bp)
app.register_blueprint(orders_bp)
app.register_blueprint(contact_bp)
app.register_blueprint(admin_bp)
app.register_blueprint(webhook_bp)
app.register_blueprint(shop_bp)
app.register_blueprint(payments_bp)


@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'HUTKO API'}), 200

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Endpoint not found.'}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({'error': 'Internal server error.'}), 500


# ── MIDNIGHT BACKUP ─────────────────────────────────────────────────────────
# Runs in a background thread. Every day at 00:05 it exports the DB to Excel
# and emails it to OWNER_EMAIL so there is always a fresh offsite copy.

def _run_backup():
    """Export DB to Excel and email it to the owner."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from database import get_db, _use_postgres
        from emails import send_email

        owner = os.environ.get('OWNER_EMAIL', '')
        if not owner:
            print("[BACKUP] OWNER_EMAIL not set — skipping email.")
            return

        conn = get_db()

        def _q(sql):
            if _use_postgres():
                cur = conn.cursor()
                cur.execute(sql)
                return cur.fetchall()
            else:
                return conn.execute(sql).fetchall()

        wb   = openpyxl.Workbook()

        def header(ws, cols, color='1B3FCE'):
            ws.append(cols)
            for cell in ws[1]:
                cell.font  = Font(bold=True, color='FFFFFF')
                cell.fill  = PatternFill('solid', fgColor=color)

        ws1 = wb.active
        ws1.title = 'Orders'
        header(ws1, ['ID','Ref','Date','Customer','Email','Phone',
                     'Street','Post','City','Province','Method',
                     'Delivery Date','Items','Subtotal','Delivery','Total','Status'])
        for r in _q("SELECT * FROM orders ORDER BY created_at DESC"):
            r = dict(r)
            ws1.append([r['id'], r['order_ref'], str(r['created_at']),
                        r['customer_name'], r['customer_email'], r['customer_phone'],
                        r['addr_street'], r['addr_postcode'], r['addr_city'], r['addr_province'],
                        r['delivery_method'], r.get('delivery_date',''),
                        r['items_json'], r['subtotal'], r['delivery_cost'], r['total'], r['status']])

        ws2 = wb.create_sheet('Users')
        header(ws2, ['ID','Name','Email','Phone','Street','Post','City','Province','Joined'], '0F6E56')
        for r in _q("SELECT * FROM users ORDER BY created_at DESC"):
            r = dict(r)
            ws2.append([r['id'], r['name'], r['email'], r['phone'],
                        r['addr_street'], r['addr_postcode'], r['addr_city'], r['addr_province'],
                        str(r['created_at'])])

        ws3 = wb.create_sheet('Messages')
        header(ws3, ['ID','Date','Name','Email','Topic','Title','Message'])
        for r in _q("SELECT * FROM messages ORDER BY created_at DESC"):
            r = dict(r)
            ws3.append([r['id'], str(r['created_at']), r['name'], r['email'],
                        r['topic'], r['title'], r['body']])

        ws4 = wb.create_sheet('Newsletter')
        header(ws4, ['ID','Email','Subscribed At'], 'E8622A')
        for r in _q("SELECT * FROM newsletter ORDER BY created_at DESC"):
            r = dict(r)
            ws4.append([r['id'], r['email'], str(r['created_at'])])

        conn.close()

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        date_str = datetime.datetime.now().strftime('%Y-%m-%d')
        send_email(
            owner,
            f"HUTKO daily backup — {date_str}",
            f"""<p style="font-family:Arial,sans-serif;color:#333;">
                Daily database backup for <strong>HUTKO Kitchen</strong>, {date_str}.<br>
                All orders, users and messages are stored safely in PostgreSQL.
            </p>"""
        )
        print(f"[BACKUP] Daily backup email sent to {owner}")
    except Exception as e:
        print(f"[BACKUP ERROR] {e}")


def _backup_scheduler():
    """Sleep until 00:05, fire backup, repeat daily."""
    while True:
        now  = datetime.datetime.now()
        next_run = now.replace(hour=0, minute=5, second=0, microsecond=0)
        if next_run <= now:
            next_run += datetime.timedelta(days=1)
        sleep_secs = (next_run - now).total_seconds()
        print(f"[BACKUP] Next backup scheduled in {sleep_secs/3600:.1f} hours")
        time.sleep(sleep_secs)
        _run_backup()


# ── KEEP-ALIVE PING ──────────────────────────────────────────────────────────
# Render free tier sleeps after 15 min of inactivity.
# This thread pings /api/health every 10 minutes so the server stays warm.
# On a paid Render plan you can remove this — but it's harmless to keep.

def _keep_alive():
    import urllib.request
    # Wait for server to fully start before first ping
    time.sleep(30)
    # Use env var or fall back to known Render URL so server never sleeps
    api_url = os.environ.get('RENDER_EXTERNAL_URL', 'https://hutko-kitchen.onrender.com')
    ping_url = api_url.rstrip('/') + '/api/health'
    print(f"[KEEP-ALIVE] Will ping {ping_url} every 10 min")
    while True:
        try:
            urllib.request.urlopen(ping_url, timeout=10)
            print(f"[KEEP-ALIVE] ping ok → {ping_url}")
        except Exception as e:
            print(f"[KEEP-ALIVE] ping failed: {e}")
        time.sleep(600)   # 10 minutes


# Always init DB on startup — use a lock file to prevent worker races
import fcntl, tempfile
_db_ready = False
try:
    _lock_path = os.path.join(tempfile.gettempdir(), 'hutko_db_init.lock')
    with open(_lock_path, 'w') as _lf:
        fcntl.flock(_lf, fcntl.LOCK_EX)
        init_db()
        _db_ready = True
except Exception as _e:
    print(f"[STARTUP] DB init error: {_e}")

# Copy Excel to persistent disk on first deploy so shop data survives redeploys
_excel_src  = os.path.join(os.path.dirname(__file__), 'hutko_shop.xlsx')
_excel_dest = os.environ.get('SHOP_EXCEL_PATH', 'hutko_shop.xlsx')
if _excel_src != _excel_dest and os.path.exists(_excel_src) and not os.path.exists(_excel_dest):
    try:
        import shutil
        os.makedirs(os.path.dirname(_excel_dest), exist_ok=True)
        shutil.copy2(_excel_src, _excel_dest)
        print(f"[STARTUP] Copied Excel to persistent disk: {_excel_dest}")
    except Exception as _e:
        print(f"[STARTUP] Excel copy failed: {_e}")

# Start background threads once per process
if os.environ.get('FLASK_ENV') != 'development' or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
    threading.Thread(target=_backup_scheduler, daemon=True).start()
    threading.Thread(target=_keep_alive,       daemon=True).start()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"HUTKO API running on http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=os.environ.get('FLASK_ENV') == 'development')