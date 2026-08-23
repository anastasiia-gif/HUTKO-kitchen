"""
HUTKO — admin.py  (v2)  Token-based admin auth, hardened.
  • Admin tokens expire (default 12h, ADMIN_TOKEN_TTL_HOURS).
  • Login lockout after repeated wrong passwords.
  • Password changeable from the UI (bcrypt hash in settings; env fallback).
  • Audit log. admin_required exported for other admin blueprints.
"""

import os, io, json, time, secrets
from functools import wraps
from datetime import datetime

import bcrypt
from flask import Blueprint, request, jsonify, send_file, g
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import get_db, _placeholder

admin_bp = Blueprint('admin', __name__)
_p = _placeholder()
TOKEN_TTL_HOURS = int(os.environ.get('ADMIN_TOKEN_TTL_HOURS', '12'))

_LOCK_MAX_FAILS = 6
_LOCK_WINDOW = 15 * 60
_fails = {}


def _hash_pw(plain):
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt()).decode()


def _stored_password_hash():
    conn = get_db()
    try:
        row = conn.execute("SELECT value FROM settings WHERE key='admin_password_hash'").fetchone()
    finally:
        conn.close()
    return (row['value'] if row and row['value'] else '')


def _check_password(plain):
    h = _stored_password_hash()
    if h:
        try:
            return bcrypt.checkpw(plain.encode(), h.encode())
        except Exception:
            return False
    env_pw = os.environ.get('ADMIN_PASSWORD', '')
    return bool(env_pw) and plain == env_pw


def _admin_configured():
    return bool(_stored_password_hash()) or bool(os.environ.get('ADMIN_PASSWORD', ''))


def _new_token():
    token = secrets.token_hex(32)
    conn = get_db()
    conn.execute(
        f"INSERT INTO admin_tokens (token, expires_at) VALUES ({_p}, datetime('now', '+{TOKEN_TTL_HOURS} hours'))",
        (token,))
    conn.commit()
    conn.close()
    return token


def _admin_token_valid(token):
    if not token:
        return False
    conn = get_db()
    row = conn.execute(
        f"SELECT token FROM admin_tokens WHERE token={_p} AND (expires_at IS NULL OR expires_at > datetime('now'))",
        (token,)).fetchone()
    conn.execute("DELETE FROM admin_tokens WHERE expires_at IS NOT NULL AND expires_at <= datetime('now')")
    conn.commit()
    conn.close()
    return row is not None


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '').strip()
        if not _admin_token_valid(token):
            return jsonify({'error': 'Admin access required.'}), 403
        g.admin_token = token
        return f(*args, **kwargs)
    return decorated


def audit(action, detail=''):
    try:
        conn = get_db()
        conn.execute(f"INSERT INTO admin_audit (actor, action, detail) VALUES ({_p},{_p},{_p})",
                     ('admin', action, detail))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[AUDIT ERROR] {e}")


def _throttled(ip):
    now = time.time()
    hits = [t for t in _fails.get(ip, []) if now - t < _LOCK_WINDOW]
    _fails[ip] = hits
    return len(hits) >= _LOCK_MAX_FAILS


def _record_fail(ip):
    _fails.setdefault(ip, []).append(time.time())


@admin_bp.route('/api/admin/login', methods=['POST'])
def admin_login():
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or 'unknown').split(',')[0].strip()
    if _throttled(ip):
        return jsonify({'error': 'Too many attempts. Please wait a few minutes and try again.'}), 429
    if not _admin_configured():
        return jsonify({'error': 'Admin access is not configured.'}), 503
    data = request.get_json(silent=True) or {}
    if not _check_password(data.get('password', '')):
        _record_fail(ip)
        return jsonify({'error': 'Wrong password.'}), 401
    _fails.pop(ip, None)
    token = _new_token()
    audit('login', f'from {ip}')
    return jsonify({'token': token, 'expires_hours': TOKEN_TTL_HOURS}), 200


@admin_bp.route('/api/admin/logout', methods=['POST'])
@admin_required
def admin_logout():
    conn = get_db()
    conn.execute(f"DELETE FROM admin_tokens WHERE token={_p}", (g.admin_token,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Logged out.'}), 200


@admin_bp.route('/api/admin/password', methods=['PUT'])
@admin_required
def admin_change_password():
    data = request.get_json(silent=True) or {}
    if not _check_password(data.get('current_password', '')):
        return jsonify({'error': 'Current password is incorrect.'}), 401
    new = data.get('new_password', '')
    if len(new) < 8:
        return jsonify({'error': 'New password must be at least 8 characters.'}), 400
    conn = get_db()
    conn.execute(
        f"INSERT INTO settings (key, value, updated_at) VALUES ('admin_password_hash', {_p}, datetime('now')) "
        f"ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=datetime('now')",
        (_hash_pw(new),))
    conn.commit()
    conn.close()
    audit('password_changed')
    return jsonify({'message': 'Password updated.'}), 200


@admin_bp.route('/api/admin/stats', methods=['GET'])
@admin_required
def stats():
    conn = get_db()

    def count(sql):
        row = conn.execute(sql).fetchone()
        return list(row)[0] if row else 0

    data = {
        'total_orders':    count("SELECT COUNT(*) FROM orders"),
        'total_revenue':   count("SELECT COALESCE(SUM(total),0) FROM orders WHERE status != 'cancelled'"),
        'total_users':     count("SELECT COUNT(*) FROM users"),
        'newsletter_subs': count("SELECT COUNT(*) FROM newsletter"),
        'unread_messages': count("SELECT COUNT(*) FROM messages"),
        'pending_orders':  count("SELECT COUNT(*) FROM orders WHERE status='confirmed'"),
        'active_products': count("SELECT COUNT(*) FROM products WHERE active=1"),
        'active_bundles':  count("SELECT COUNT(*) FROM bundles WHERE active=1"),
    }
    conn.close()
    return jsonify(data), 200


@admin_bp.route('/api/admin/export', methods=['GET'])
@admin_required
def export_excel():
    conn = get_db()
    wb = Workbook()

    def header(ws, cols, color="1a2356"):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color="FFFFFF", name="Calibri")
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center")

    ws1 = wb.active
    ws1.title = "Orders"
    header(ws1, ["Order Ref", "Date", "Customer", "Email", "Phone", "Address", "City",
                 "Province", "Method", "Delivery Date", "Items", "Subtotal", "Delivery", "Total", "Status"])
    for r in conn.execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall():
        r = dict(r)
        try:
            items_str = " | ".join(f"{i['name']} x{i['qty']}" for i in json.loads(r['items_json']))
        except Exception:
            items_str = r.get('items_json', '')
        ws1.append([r['order_ref'], str(r['created_at']), r['customer_name'], r['customer_email'],
                    r['customer_phone'], f"{r['addr_street']}, {r['addr_postcode']}", r['addr_city'],
                    r['addr_province'], r['delivery_method'], r.get('delivery_date', ''), items_str,
                    r['subtotal'], r['delivery_cost'], r['total'], r['status']])

    ws2 = wb.create_sheet("Users")
    header(ws2, ["ID", "Name", "Email", "Phone", "Street", "Postcode", "City", "Province", "Registered"], "E8622A")
    for r in conn.execute("SELECT * FROM users ORDER BY created_at DESC").fetchall():
        ws2.append([r['id'], r['name'], r['email'], r['phone'], r['addr_street'],
                    r['addr_postcode'], r['addr_city'], r['addr_province'], str(r['created_at'])])

    ws3 = wb.create_sheet("Messages")
    header(ws3, ["ID", "Date", "Name", "Email", "Topic", "Title", "Message"])
    for r in conn.execute("SELECT * FROM messages ORDER BY created_at DESC").fetchall():
        ws3.append([r['id'], str(r['created_at']), r['name'], r['email'], r['topic'], r['title'], r['body']])

    ws4 = wb.create_sheet("Newsletter")
    header(ws4, ["ID", "Email", "Subscribed At"], "E8622A")
    for r in conn.execute("SELECT * FROM newsletter ORDER BY created_at DESC").fetchall():
        ws4.append([r['id'], r['email'], str(r['created_at'])])

    conn.close()
    for ws in wb.worksheets:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or '')) for c in col) + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"hutko_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@admin_bp.route('/api/admin/audit', methods=['GET'])
@admin_required
def audit_log():
    conn = get_db()
    rows = conn.execute("SELECT actor, action, detail, created_at FROM admin_audit ORDER BY id DESC LIMIT 100").fetchall()
    conn.close()
    return jsonify({'events': [dict(r) for r in rows]}), 200
