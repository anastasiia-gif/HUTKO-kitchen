"""
HUTKO — admin_catalog.py
Admin CRUD for products & bundles, image upload, Excel import/export.
All write endpoints require the admin token.
"""

import os, io, json, re, time, secrets
from flask import Blueprint, request, jsonify, send_file, send_from_directory
from database import get_db, _placeholder
from admin import admin_required, audit

catalog_bp = Blueprint('admin_catalog', __name__)
_p = _placeholder()

MEDIA_DIR = os.environ.get('MEDIA_PATH', os.path.join(os.path.dirname(os.path.abspath(__file__)), 'media'))
os.makedirs(MEDIA_DIR, exist_ok=True)
ALLOWED_IMG = {'.png', '.jpg', '.jpeg', '.webp', '.gif'}

PRODUCT_FIELDS = [
    'category', 'name_en', 'name_ua', 'name_nl', 'desc_en', 'desc_ua', 'desc_nl',
    'about_en', 'about_ua', 'about_nl', 'prepare_en', 'prepare_ua', 'prepare_nl',
    'ingredients_en', 'ingredients_ua', 'ingredients_nl', 'hutko_tip_en', 'hutko_tip_ua',
    'hutko_tip_nl', 'storage_en', 'storage_ua', 'storage_nl', 'unit', 'badge', 'photo',
]
BUNDLE_FIELDS = ['name_en', 'name_ua', 'name_nl', 'size_label', 'photo', 'badge',
                 'choice_en', 'choice_ua', 'choice_nl']


def _slugify(s):
    s = re.sub(r'[^a-z0-9]+', '-', str(s).lower()).strip('-')
    return s or ('item-' + secrets.token_hex(3))


def _as_json_list(val):
    if isinstance(val, list):
        return json.dumps(val)
    if val is None:
        return json.dumps([])
    return json.dumps([x.strip() for x in str(val).split(',') if x.strip()])


def _num(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _product_row(conn, row):
    d = dict(row)
    vrows = conn.execute(
        f"SELECT label, price, active FROM product_variants WHERE product_id={_p} ORDER BY sort_order, id",
        (d['id'],)).fetchall()
    d['variants'] = [dict(v) for v in vrows]
    for k in ('gallery', 'dietary'):
        try:
            d[k] = json.loads(d.get(k) or '[]')
        except (ValueError, TypeError):
            d[k] = []
    return d


@catalog_bp.route('/api/admin/products', methods=['GET'])
@admin_required
def list_products():
    conn = get_db()
    rows = conn.execute("SELECT * FROM products ORDER BY sort_order, name_en").fetchall()
    out = [_product_row(conn, r) for r in rows]
    conn.close()
    return jsonify({'products': out}), 200


@catalog_bp.route('/api/admin/products', methods=['POST'])
@admin_required
def create_product():
    data = request.get_json(silent=True) or {}
    pid = _slugify(data.get('id') or data.get('name_en') or '')
    conn = get_db()
    if conn.execute(f"SELECT id FROM products WHERE id={_p}", (pid,)).fetchone():
        conn.close()
        return jsonify({'error': f'A product with id "{pid}" already exists.'}), 409
    cols = ['id'] + PRODUCT_FIELDS + ['base_price', 'gallery', 'dietary', 'active', 'sort_order']
    vals = [pid] + [str(data.get(f, '') or '') for f in PRODUCT_FIELDS] + [
        _num(data.get('base_price')), _as_json_list(data.get('gallery')), _as_json_list(data.get('dietary')),
        1 if data.get('active', True) else 0, int(data.get('sort_order') or 0)]
    conn.execute(f"INSERT INTO products ({','.join(cols)}) VALUES ({','.join([_p]*len(cols))})", vals)
    _save_variants(conn, pid, data.get('variants', []))
    conn.commit()
    conn.close()
    audit('product_create', pid)
    return jsonify({'id': pid, 'message': 'Product created.'}), 201


@catalog_bp.route('/api/admin/products/<pid>', methods=['PUT'])
@admin_required
def update_product(pid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    if not conn.execute(f"SELECT id FROM products WHERE id={_p}", (pid,)).fetchone():
        conn.close()
        return jsonify({'error': 'Product not found.'}), 404
    sets, vals = [], []
    for f in PRODUCT_FIELDS:
        if f in data:
            sets.append(f"{f}={_p}"); vals.append(str(data.get(f) or ''))
    if 'base_price' in data:
        sets.append(f"base_price={_p}"); vals.append(_num(data.get('base_price')))
    if 'gallery' in data:
        sets.append(f"gallery={_p}"); vals.append(_as_json_list(data.get('gallery')))
    if 'dietary' in data:
        sets.append(f"dietary={_p}"); vals.append(_as_json_list(data.get('dietary')))
    if 'active' in data:
        sets.append(f"active={_p}"); vals.append(1 if data.get('active') else 0)
    if 'sort_order' in data:
        sets.append(f"sort_order={_p}"); vals.append(int(data.get('sort_order') or 0))
    sets.append("updated_at=datetime('now')")
    if sets:
        conn.execute(f"UPDATE products SET {','.join(sets)} WHERE id={_p}", (*vals, pid))
    if 'variants' in data:
        conn.execute(f"DELETE FROM product_variants WHERE product_id={_p}", (pid,))
        _save_variants(conn, pid, data.get('variants', []))
    conn.commit()
    conn.close()
    audit('product_update', pid)
    return jsonify({'message': 'Product updated.'}), 200


@catalog_bp.route('/api/admin/products/<pid>', methods=['DELETE'])
@admin_required
def delete_product(pid):
    conn = get_db()
    conn.execute(f"DELETE FROM product_variants WHERE product_id={_p}", (pid,))
    conn.execute(f"DELETE FROM products WHERE id={_p}", (pid,))
    conn.commit()
    conn.close()
    audit('product_delete', pid)
    return jsonify({'message': 'Product deleted.'}), 200


def _save_variants(conn, pid, variants):
    for i, v in enumerate(variants or []):
        conn.execute(
            f"INSERT INTO product_variants (product_id, label, price, active, sort_order) VALUES ({_p},{_p},{_p},{_p},{_p})",
            (pid, str(v.get('label', '') or ''), _num(v.get('price')), 1 if v.get('active', True) else 0, i))


@catalog_bp.route('/api/admin/bundles', methods=['GET'])
@admin_required
def list_bundles():
    conn = get_db()
    rows = conn.execute("SELECT * FROM bundles ORDER BY sort_order, name_en").fetchall()
    out = []
    for r in rows:
        d = dict(r)
        try:
            d['items'] = json.loads(d.get('items') or '[]')
        except (ValueError, TypeError):
            d['items'] = []
        out.append(d)
    conn.close()
    return jsonify({'bundles': out}), 200


@catalog_bp.route('/api/admin/bundles', methods=['POST'])
@admin_required
def create_bundle():
    data = request.get_json(silent=True) or {}
    bid = _slugify(data.get('id') or data.get('name_en') or '')
    conn = get_db()
    if conn.execute(f"SELECT id FROM bundles WHERE id={_p}", (bid,)).fetchone():
        conn.close()
        return jsonify({'error': f'A bundle with id "{bid}" already exists.'}), 409
    cols = ['id'] + BUNDLE_FIELDS + ['items', 'original_price', 'discount_price', 'active', 'sort_order']
    vals = [bid] + [str(data.get(f, '') or '') for f in BUNDLE_FIELDS] + [
        json.dumps(data.get('items') or []), _num(data.get('original_price')), _num(data.get('discount_price')),
        1 if data.get('active', True) else 0, int(data.get('sort_order') or 0)]
    conn.execute(f"INSERT INTO bundles ({','.join(cols)}) VALUES ({','.join([_p]*len(cols))})", vals)
    conn.commit()
    conn.close()
    audit('bundle_create', bid)
    return jsonify({'id': bid, 'message': 'Bundle created.'}), 201


@catalog_bp.route('/api/admin/bundles/<bid>', methods=['PUT'])
@admin_required
def update_bundle(bid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    if not conn.execute(f"SELECT id FROM bundles WHERE id={_p}", (bid,)).fetchone():
        conn.close()
        return jsonify({'error': 'Bundle not found.'}), 404
    sets, vals = [], []
    for f in BUNDLE_FIELDS:
        if f in data:
            sets.append(f"{f}={_p}"); vals.append(str(data.get(f) or ''))
    if 'items' in data:
        sets.append(f"items={_p}"); vals.append(json.dumps(data.get('items') or []))
    for nf in ('original_price', 'discount_price'):
        if nf in data:
            sets.append(f"{nf}={_p}"); vals.append(_num(data.get(nf)))
    if 'active' in data:
        sets.append(f"active={_p}"); vals.append(1 if data.get('active') else 0)
    if 'sort_order' in data:
        sets.append(f"sort_order={_p}"); vals.append(int(data.get('sort_order') or 0))
    sets.append("updated_at=datetime('now')")
    conn.execute(f"UPDATE bundles SET {','.join(sets)} WHERE id={_p}", (*vals, bid))
    conn.commit()
    conn.close()
    audit('bundle_update', bid)
    return jsonify({'message': 'Bundle updated.'}), 200


@catalog_bp.route('/api/admin/bundles/<bid>', methods=['DELETE'])
@admin_required
def delete_bundle(bid):
    conn = get_db()
    conn.execute(f"DELETE FROM bundles WHERE id={_p}", (bid,))
    conn.commit()
    conn.close()
    audit('bundle_delete', bid)
    return jsonify({'message': 'Bundle deleted.'}), 200


@catalog_bp.route('/api/admin/media', methods=['POST'])
@admin_required
def upload_media():
    if 'file' not in request.files:
        return jsonify({'error': 'No file. Use form field "file".'}), 400
    f = request.files['file']
    ext = os.path.splitext(f.filename or '')[1].lower()
    if ext not in ALLOWED_IMG:
        return jsonify({'error': f'Unsupported type {ext}. Use PNG/JPG/WEBP.'}), 400
    name = f"{_slugify(os.path.splitext(f.filename)[0])}-{int(time.time())}{ext}"
    path = os.path.join(MEDIA_DIR, name)
    f.save(path)
    try:
        from PIL import Image
        img = Image.open(path)
        if img.width > 1400:
            ratio = 1400 / img.width
            img = img.resize((1400, int(img.height * ratio)))
            img.save(path)
    except Exception:
        pass
    audit('media_upload', name)
    return jsonify({'url': f'/api/media/{name}', 'filename': name}), 201


@catalog_bp.route('/api/media/<path:filename>', methods=['GET'])
def serve_media(filename):
    return send_from_directory(MEDIA_DIR, filename)


@catalog_bp.route('/api/admin/export-catalogue', methods=['GET'])
@admin_required
def export_catalogue():
    from openpyxl import Workbook
    conn = get_db()
    wb = Workbook()
    ws = wb.active; ws.title = 'Products'
    ws.append([None])
    prod_cols = ['id', 'category', 'name_en', 'name_ua', 'name_nl',
                 'description_en', 'description_ua', 'description_nl',
                 'about_en', 'about_ua', 'about_nl', 'prepare_en', 'prepare_ua', 'prepare_nl',
                 'ingredients_en', 'ingredients_ua', 'ingredients_nl',
                 'hutko_tip_en', 'hutko_tip_ua', 'hutko_tip_nl',
                 'storage_en', 'storage_ua', 'storage_nl',
                 'base_price', 'unit', 'badge', 'photo_file', 'gallery', 'dietary', 'active']
    ws.append([None] + prod_cols)
    for r in conn.execute("SELECT * FROM products ORDER BY sort_order, name_en").fetchall():
        r = dict(r)
        try:
            gallery = ','.join(json.loads(r.get('gallery') or '[]'))
            dietary = ','.join(json.loads(r.get('dietary') or '[]'))
        except (ValueError, TypeError):
            gallery, dietary = '', ''
        ws.append([None, r['id'], r['category'], r['name_en'], r['name_ua'], r['name_nl'],
                   r['desc_en'], r['desc_ua'], r['desc_nl'], r['about_en'], r['about_ua'], r['about_nl'],
                   r['prepare_en'], r['prepare_ua'], r['prepare_nl'],
                   r['ingredients_en'], r['ingredients_ua'], r['ingredients_nl'],
                   r['hutko_tip_en'], r['hutko_tip_ua'], r['hutko_tip_nl'],
                   r['storage_en'], r['storage_ua'], r['storage_nl'],
                   r['base_price'], r['unit'], r['badge'], r['photo'], gallery, dietary,
                   'yes' if r['active'] else 'no'])
    ws.append([None, 'product_id', 'label', 'price', 'active'])
    for v in conn.execute("SELECT * FROM product_variants ORDER BY product_id, sort_order").fetchall():
        ws.append([None, v['product_id'], v['label'], v['price'], 'yes' if v['active'] else 'no'])

    wb2 = wb.create_sheet('Bundles')
    wb2.append([None])
    bcols = ['id', 'name_en', 'name_ua', 'name_nl', 'size_label', 'items',
             'original_price', 'discount_price', 'photo_file', 'badge', 'choice_en', 'choice_ua', 'choice_nl']
    wb2.append([None] + bcols)
    for r in conn.execute("SELECT * FROM bundles ORDER BY sort_order, name_en").fetchall():
        r = dict(r)
        try:
            items = ','.join(f"{i['product_id']}:{i['qty']}" for i in json.loads(r.get('items') or '[]'))
        except (ValueError, TypeError):
            items = ''
        wb2.append([None, r['id'], r['name_en'], r['name_ua'], r['name_nl'], r['size_label'], items,
                    r['original_price'], r['discount_price'], r['photo'], r['badge'],
                    r['choice_en'], r['choice_ua'], r['choice_nl']])

    ws3 = wb.create_sheet('Settings')
    ws3.append(['setting', 'value'])
    for r in conn.execute("SELECT key, value FROM settings ORDER BY key").fetchall():
        ws3.append([r['key'], r['value']])
    conn.close()
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='hutko_catalogue.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@catalog_bp.route('/api/admin/import-catalogue', methods=['POST'])
@admin_required
def import_catalogue():
    if 'file' not in request.files:
        return jsonify({'error': 'No file. Use form field "file".'}), 400
    f = request.files['file']
    if not (f.filename or '').endswith('.xlsx'):
        return jsonify({'error': 'Only .xlsx accepted.'}), 400
    tmp = os.path.join(MEDIA_DIR, f'_import_{int(time.time())}.xlsx')
    f.save(tmp)
    try:
        from migrate_excel_to_db import parse_workbook, import_into_db
        summary = import_into_db(parse_workbook(tmp))
        audit('catalogue_import', json.dumps(summary))
        return jsonify({'message': 'Imported.', 'summary': summary}), 200
    except Exception as e:
        return jsonify({'error': f'Import failed: {e}'}), 500
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass
